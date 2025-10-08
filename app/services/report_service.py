from sqlalchemy.orm import Session
from app.services.analytics_service import AnalyticsService
from app.schemas.analytics import (
    ReportRequest, ReportResponse, ExportRequest, ExportResponse,
    ReportType, ReportFormat, AnalyticsPeriod
)
from typing import Optional, List, Dict, Any
import structlog
from datetime import datetime, timedelta
import json
import csv
import io
import uuid
from pathlib import Path

logger = structlog.get_logger()

class ReportService:
    """Servicio para generación de reportes y exportación de datos"""
    
    def __init__(self, db: Session):
        self.db = db
        self.analytics_service = AnalyticsService(db)
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
    
    async def generate_report(self, report_request: ReportRequest) -> ReportResponse:
        """Generar reporte"""
        try:
            report_id = str(uuid.uuid4())
            
            # Calcular fechas
            end_date = datetime.utcnow()
            if report_request.period == "custom" and report_request.start_date and report_request.end_date:
                start_date = report_request.start_date
                end_date = report_request.end_date
            else:
                start_date = self._calculate_start_date(report_request.period, end_date)
            
            # Generar datos del reporte
            report_data = await self._generate_report_data(
                report_request, start_date, end_date
            )
            
            # Generar archivo según formato
            file_path = await self._generate_report_file(
                report_id, report_request, report_data
            )
            
            # Calcular tamaño del archivo
            file_size = file_path.stat().st_size if file_path.exists() else 0
            
            # Fecha de expiración (7 días)
            expires_at = datetime.utcnow() + timedelta(days=7)
            
            response = ReportResponse(
                report_id=report_id,
                report_type=report_request.report_type,
                format=report_request.format,
                status="completed",
                download_url=f"/reports/{report_id}",
                file_size=file_size,
                generated_at=datetime.utcnow(),
                expires_at=expires_at
            )
            
            logger.info("Report generated", 
                       report_id=report_id, 
                       report_type=report_request.report_type.value,
                       format=report_request.format.value,
                       file_size=file_size)
            
            return response
            
        except Exception as e:
            logger.error("Error generating report", 
                        report_type=report_request.report_type.value, 
                        error=str(e))
            return ReportResponse(
                report_id=str(uuid.uuid4()),
                report_type=report_request.report_type,
                format=report_request.format,
                status="failed",
                error_message=str(e),
                generated_at=datetime.utcnow()
            )
    
    async def _generate_report_data(
        self, 
        report_request: ReportRequest, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Generar datos del reporte"""
        try:
            report_data = {
                "metadata": {
                    "site_id": report_request.site_id,
                    "report_type": report_request.report_type.value,
                    "period": report_request.period.value,
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "generated_at": datetime.utcnow().isoformat()
                }
            }
            
            # Generar datos según tipo de reporte
            if report_request.report_type == ReportType.USER_ACTIVITY:
                report_data["data"] = await self._get_user_activity_data(
                    report_request.site_id, start_date, end_date
                )
            elif report_request.report_type == ReportType.ENGAGEMENT:
                report_data["data"] = await self._get_engagement_data(
                    report_request.site_id, start_date, end_date
                )
            elif report_request.report_type == ReportType.STICKER_USAGE:
                report_data["data"] = await self._get_sticker_usage_data(
                    report_request.site_id, start_date, end_date
                )
            elif report_request.report_type == ReportType.VIDEO_ANALYTICS:
                report_data["data"] = await self._get_video_analytics_data(
                    report_request.site_id, start_date, end_date
                )
            elif report_request.report_type == ReportType.NOTIFICATION_PERFORMANCE:
                report_data["data"] = await self._get_notification_performance_data(
                    report_request.site_id, start_date, end_date
                )
            else:
                # Reporte personalizado
                report_data["data"] = await self._get_custom_report_data(
                    report_request, start_date, end_date
                )
            
            return report_data
            
        except Exception as e:
            logger.error("Error generating report data", 
                        report_type=report_request.report_type.value, 
                        error=str(e))
            return {"error": str(e)}
    
    async def _get_user_activity_data(
        self, 
        site_id: str, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Obtener datos de actividad de usuarios"""
        try:
            from app.models.user import User
            from app.models.interaction import Interaction
            
            # Usuarios totales
            total_users = self.db.query(User).filter(
                User.site_id == site_id,
                User.activo == True
            ).count()
            
            # Usuarios activos en el período
            active_users = self.db.query(User).join(Interaction).filter(
                User.site_id == site_id,
                User.activo == True,
                Interaction.site_id == site_id,
                Interaction.fecha_interaccion >= start_date,
                Interaction.fecha_interaccion <= end_date
            ).distinct().count()
            
            # Usuarios nuevos
            new_users = self.db.query(User).filter(
                User.site_id == site_id,
                User.activo == True,
                User.fecha_registro >= start_date,
                User.fecha_registro <= end_date
            ).count()
            
            # Usuarios por mes de registro
            users_by_month = []
            current_date = start_date
            while current_date <= end_date:
                month_start = current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                if current_date.month == 12:
                    month_end = current_date.replace(year=current_date.year + 1, month=1, day=1)
                else:
                    month_end = current_date.replace(month=current_date.month + 1, day=1)
                
                count = self.db.query(User).filter(
                    User.site_id == site_id,
                    User.activo == True,
                    User.fecha_registro >= month_start,
                    User.fecha_registro < month_end
                ).count()
                
                users_by_month.append({
                    "month": month_start.strftime("%Y-%m"),
                    "count": count
                })
                
                current_date = month_end
            
            return {
                "total_users": total_users,
                "active_users": active_users,
                "new_users": new_users,
                "users_by_month": users_by_month
            }
            
        except Exception as e:
            logger.error("Error getting user activity data", 
                        site_id=site_id, 
                        error=str(e))
            return {}
    
    async def _get_engagement_data(
        self, 
        site_id: str, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Obtener datos de engagement"""
        try:
            from app.models.interaction import Interaction
            
            # Total de interacciones
            total_interactions = self.db.query(Interaction).filter(
                Interaction.site_id == site_id,
                Interaction.fecha_interaccion >= start_date,
                Interaction.fecha_interaccion <= end_date
            ).count()
            
            # Interacciones por tipo
            interactions_by_type = []
            interaction_types = ['like', 'comment', 'review', 'share', 'view']
            
            for interaction_type in interaction_types:
                count = self.db.query(Interaction).filter(
                    Interaction.site_id == site_id,
                    Interaction.tipo_interaccion == interaction_type,
                    Interaction.fecha_interaccion >= start_date,
                    Interaction.fecha_interaccion <= end_date
                ).count()
                
                interactions_by_type.append({
                    "type": interaction_type,
                    "count": count,
                    "percentage": (count / total_interactions * 100) if total_interactions > 0 else 0
                })
            
            # Interacciones diarias
            daily_interactions = []
            current_date = start_date.date()
            end_date_only = end_date.date()
            
            while current_date <= end_date_only:
                day_start = datetime.combine(current_date, datetime.min.time())
                day_end = datetime.combine(current_date, datetime.max.time())
                
                count = self.db.query(Interaction).filter(
                    Interaction.site_id == site_id,
                    Interaction.fecha_interaccion >= day_start,
                    Interaction.fecha_interaccion <= day_end
                ).count()
                
                daily_interactions.append({
                    "date": current_date.isoformat(),
                    "count": count
                })
                
                current_date += timedelta(days=1)
            
            return {
                "total_interactions": total_interactions,
                "interactions_by_type": interactions_by_type,
                "daily_interactions": daily_interactions
            }
            
        except Exception as e:
            logger.error("Error getting engagement data", 
                        site_id=site_id, 
                        error=str(e))
            return {}
    
    async def _get_sticker_usage_data(
        self, 
        site_id: str, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Obtener datos de uso de stickers"""
        try:
            from app.models.sticker import Sticker
            
            # Total de stickers generados
            total_generated = self.db.query(Sticker).filter(
                Sticker.site_id == site_id,
                Sticker.fecha_generacion >= start_date,
                Sticker.fecha_generacion <= end_date
            ).count()
            
            # Total de stickers usados
            total_used = self.db.query(Sticker).filter(
                Sticker.site_id == site_id,
                Sticker.usado == True,
                Sticker.fecha_uso >= start_date,
                Sticker.fecha_uso <= end_date
            ).count()
            
            # Stickers por tipo
            stickers_by_type = []
            sticker_types = ['registro', 'instagram', 'reseña', 'video_completado']
            
            for sticker_type in sticker_types:
                generated = self.db.query(Sticker).filter(
                    Sticker.site_id == site_id,
                    Sticker.tipo_sticker == sticker_type,
                    Sticker.fecha_generacion >= start_date,
                    Sticker.fecha_generacion <= end_date
                ).count()
                
                used = self.db.query(Sticker).filter(
                    Sticker.site_id == site_id,
                    Sticker.tipo_sticker == sticker_type,
                    Sticker.usado == True,
                    Sticker.fecha_uso >= start_date,
                    Sticker.fecha_uso <= end_date
                ).count()
                
                stickers_by_type.append({
                    "type": sticker_type,
                    "generated": generated,
                    "used": used,
                    "usage_rate": (used / generated * 100) if generated > 0 else 0
                })
            
            return {
                "total_generated": total_generated,
                "total_used": total_used,
                "usage_rate": (total_used / total_generated * 100) if total_generated > 0 else 0,
                "stickers_by_type": stickers_by_type
            }
            
        except Exception as e:
            logger.error("Error getting sticker usage data", 
                        site_id=site_id, 
                        error=str(e))
            return {}
    
    async def _get_video_analytics_data(
        self, 
        site_id: str, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Obtener datos de analytics de videos"""
        try:
            from app.models.video import Video, VideoWatchSession, VideoCompletion
            
            # Total de videos
            total_videos = self.db.query(Video).filter(
                Video.site_id == site_id,
                Video.activo == True
            ).count()
            
            # Total de visualizaciones
            total_views = self.db.query(VideoWatchSession).join(Video).filter(
                Video.site_id == site_id,
                VideoWatchSession.start_time >= start_date,
                VideoWatchSession.start_time <= end_date
            ).count()
            
            # Total de completaciones
            total_completions = self.db.query(VideoCompletion).join(Video).filter(
                Video.site_id == site_id,
                VideoCompletion.completed_at >= start_date,
                VideoCompletion.completed_at <= end_date
            ).count()
            
            # Videos más vistos
            most_watched = self.db.query(
                Video.id,
                Video.titulo,
                func.count(VideoWatchSession.id).label('view_count')
            ).join(VideoWatchSession).filter(
                Video.site_id == site_id,
                VideoWatchSession.start_time >= start_date,
                VideoWatchSession.start_time <= end_date
            ).group_by(Video.id, Video.titulo).order_by(desc('view_count')).limit(10).all()
            
            most_watched_list = []
            for video in most_watched:
                most_watched_list.append({
                    "video_id": video.id,
                    "title": video.titulo,
                    "view_count": video.view_count
                })
            
            return {
                "total_videos": total_videos,
                "total_views": total_views,
                "total_completions": total_completions,
                "completion_rate": (total_completions / total_views * 100) if total_views > 0 else 0,
                "most_watched_videos": most_watched_list
            }
            
        except Exception as e:
            logger.error("Error getting video analytics data", 
                        site_id=site_id, 
                        error=str(e))
            return {}
    
    async def _get_notification_performance_data(
        self, 
        site_id: str, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Obtener datos de rendimiento de notificaciones"""
        try:
            from app.models.notification import Notification
            
            # Total de notificaciones enviadas
            total_sent = self.db.query(Notification).filter(
                Notification.site_id == site_id,
                Notification.created_at >= start_date,
                Notification.created_at <= end_date,
                Notification.status == 'sent'
            ).count()
            
            # Total de notificaciones entregadas
            total_delivered = self.db.query(Notification).filter(
                Notification.site_id == site_id,
                Notification.created_at >= start_date,
                Notification.created_at <= end_date,
                Notification.delivered_at.isnot(None)
            ).count()
            
            # Total de notificaciones leídas
            total_read = self.db.query(Notification).filter(
                Notification.site_id == site_id,
                Notification.created_at >= start_date,
                Notification.created_at <= end_date,
                Notification.read_at.isnot(None)
            ).count()
            
            # Notificaciones por tipo
            notifications_by_type = []
            notification_types = ['sticker', 'instagram', 'points', 'level_up', 'system', 'video_completed']
            
            for notification_type in notification_types:
                count = self.db.query(Notification).filter(
                    Notification.site_id == site_id,
                    Notification.type == notification_type,
                    Notification.created_at >= start_date,
                    Notification.created_at <= end_date
                ).count()
                
                notifications_by_type.append({
                    "type": notification_type,
                    "count": count,
                    "percentage": (count / total_sent * 100) if total_sent > 0 else 0
                })
            
            return {
                "total_sent": total_sent,
                "total_delivered": total_delivered,
                "total_read": total_read,
                "delivery_rate": (total_delivered / total_sent * 100) if total_sent > 0 else 0,
                "read_rate": (total_read / total_delivered * 100) if total_delivered > 0 else 0,
                "notifications_by_type": notifications_by_type
            }
            
        except Exception as e:
            logger.error("Error getting notification performance data", 
                        site_id=site_id, 
                        error=str(e))
            return {}
    
    async def _get_custom_report_data(
        self, 
        report_request: ReportRequest, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Any]:
        """Obtener datos de reporte personalizado"""
        try:
            # Usar el servicio de analytics para obtener datos personalizados
            from app.schemas.analytics import AnalyticsQuery
            
            query = AnalyticsQuery(
                site_id=report_request.site_id,
                period=report_request.period,
                start_date=start_date,
                end_date=end_date,
                metrics=report_request.metrics,
                dimensions=report_request.dimensions,
                filters=report_request.filters
            )
            
            analytics_response = await self.analytics_service.get_custom_analytics(query)
            
            return {
                "metrics": [metric.dict() for metric in analytics_response.metrics],
                "dimensions": [dim.dict() for dim in analytics_response.dimensions] if analytics_response.dimensions else [],
                "time_series": [ts.dict() for ts in analytics_response.time_series] if analytics_response.time_series else []
            }
            
        except Exception as e:
            logger.error("Error getting custom report data", 
                        error=str(e))
            return {}
    
    async def _generate_report_file(
        self, 
        report_id: str, 
        report_request: ReportRequest, 
        report_data: Dict[str, Any]
    ) -> Path:
        """Generar archivo del reporte"""
        try:
            file_path = self.reports_dir / f"{report_id}.{report_request.format.value}"
            
            if report_request.format == ReportFormat.JSON:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(report_data, f, indent=2, ensure_ascii=False, default=str)
            
            elif report_request.format == ReportFormat.CSV:
                await self._generate_csv_report(file_path, report_data)
            
            elif report_request.format == ReportFormat.PDF:
                await self._generate_pdf_report(file_path, report_data, report_request)
            
            elif report_request.format == ReportFormat.XLSX:
                await self._generate_xlsx_report(file_path, report_data)
            
            elif report_request.format == ReportFormat.HTML:
                await self._generate_html_report(file_path, report_data, report_request)
            
            return file_path
            
        except Exception as e:
            logger.error("Error generating report file", 
                        report_id=report_id, 
                        format=report_request.format.value, 
                        error=str(e))
            raise
    
    async def _generate_csv_report(self, file_path: Path, report_data: Dict[str, Any]):
        """Generar reporte CSV"""
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Escribir metadatos
                writer.writerow(['Metadata'])
                for key, value in report_data.get('metadata', {}).items():
                    writer.writerow([key, value])
                
                writer.writerow([])  # Línea en blanco
                
                # Escribir datos
                data = report_data.get('data', {})
                if isinstance(data, dict):
                    writer.writerow(['Data'])
                    for key, value in data.items():
                        if isinstance(value, list):
                            writer.writerow([key])
                            for item in value:
                                if isinstance(item, dict):
                                    writer.writerow([item.get('type', ''), item.get('count', 0)])
                                else:
                                    writer.writerow([item])
                        else:
                            writer.writerow([key, value])
                
        except Exception as e:
            logger.error("Error generating CSV report", error=str(e))
            raise
    
    async def _generate_pdf_report(self, file_path: Path, report_data: Dict[str, Any], report_request: ReportRequest):
        """Generar reporte PDF"""
        try:
            # Implementar generación de PDF usando reportlab
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            
            doc = SimpleDocTemplate(str(file_path), pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title = Paragraph(f"Reporte: {report_request.report_type.value}", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Metadatos
            metadata = report_data.get('metadata', {})
            for key, value in metadata.items():
                para = Paragraph(f"<b>{key}:</b> {value}", styles['Normal'])
                story.append(para)
            
            story.append(Spacer(1, 12))
            
            # Datos
            data = report_data.get('data', {})
            if isinstance(data, dict):
                for key, value in data.items():
                    para = Paragraph(f"<b>{key}:</b> {value}", styles['Normal'])
                    story.append(para)
            
            doc.build(story)
            
        except Exception as e:
            logger.error("Error generating PDF report", error=str(e))
            raise
    
    async def _generate_xlsx_report(self, file_path: Path, report_data: Dict[str, Any]):
        """Generar reporte XLSX"""
        try:
            # Implementar generación de XLSX usando openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            # Escribir metadatos
            row = 1
            ws.cell(row=row, column=1, value="Metadata")
            row += 1
            
            metadata = report_data.get('metadata', {})
            for key, value in metadata.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            row += 1  # Línea en blanco
            
            # Escribir datos
            data = report_data.get('data', {})
            if isinstance(data, dict):
                ws.cell(row=row, column=1, value="Data")
                row += 1
                
                for key, value in data.items():
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error("Error generating XLSX report", error=str(e))
            raise
    
    async def _generate_html_report(self, file_path: Path, report_data: Dict[str, Any], report_request: ReportRequest):
        """Generar reporte HTML"""
        try:
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Reporte: {report_request.report_type.value}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #333; }}
                    h2 {{ color: #666; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f2f2f2; }}
                </style>
            </head>
            <body>
                <h1>Reporte: {report_request.report_type.value}</h1>
                
                <h2>Metadatos</h2>
                <table>
                    <tr><th>Campo</th><th>Valor</th></tr>
            """
            
            metadata = report_data.get('metadata', {})
            for key, value in metadata.items():
                html_content += f"<tr><td>{key}</td><td>{value}</td></tr>"
            
            html_content += """
                </table>
                
                <h2>Datos</h2>
                <table>
                    <tr><th>Campo</th><th>Valor</th></tr>
            """
            
            data = report_data.get('data', {})
            if isinstance(data, dict):
                for key, value in data.items():
                    html_content += f"<tr><td>{key}</td><td>{value}</td></tr>"
            
            html_content += """
                </table>
            </body>
            </html>
            """
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
        except Exception as e:
            logger.error("Error generating HTML report", error=str(e))
            raise
    
    def _calculate_start_date(self, period: str, end_date: datetime) -> datetime:
        """Calcular fecha de inicio según el período"""
        if period.endswith('d'):
            days = int(period[:-1])
            return end_date - timedelta(days=days)
        elif period.endswith('w'):
            weeks = int(period[:-1])
            return end_date - timedelta(weeks=weeks)
        elif period.endswith('m'):
            months = int(period[:-1])
            return end_date - timedelta(days=months * 30)
        else:
            return end_date - timedelta(days=30)
    
    async def export_data(self, export_request: ExportRequest) -> ExportResponse:
        """Exportar datos"""
        try:
            export_id = str(uuid.uuid4())
            
            # Generar datos de exportación
            export_data = await self._get_export_data(export_request)
            
            # Generar archivo
            file_path = await self._generate_export_file(export_id, export_request, export_data)
            
            # Calcular tamaño del archivo
            file_size = file_path.stat().st_size if file_path.exists() else 0
            record_count = len(export_data) if isinstance(export_data, list) else 1
            
            # Fecha de expiración (7 días)
            expires_at = datetime.utcnow() + timedelta(days=7)
            
            response = ExportResponse(
                export_id=export_id,
                data_type=export_request.data_type,
                format=export_request.format,
                status="completed",
                download_url=f"/exports/{export_id}",
                file_size=file_size,
                record_count=record_count,
                generated_at=datetime.utcnow(),
                expires_at=expires_at
            )
            
            logger.info("Data exported", 
                       export_id=export_id, 
                       data_type=export_request.data_type,
                       format=export_request.format.value,
                       record_count=record_count)
            
            return response
            
        except Exception as e:
            logger.error("Error exporting data", 
                        data_type=export_request.data_type, 
                        error=str(e))
            return ExportResponse(
                export_id=str(uuid.uuid4()),
                data_type=export_request.data_type,
                format=export_request.format,
                status="failed",
                error_message=str(e),
                generated_at=datetime.utcnow()
            )
    
    async def _get_export_data(self, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Obtener datos para exportación"""
        try:
            if export_request.data_type == "users":
                return await self._export_users_data(export_request)
            elif export_request.data_type == "stickers":
                return await self._export_stickers_data(export_request)
            elif export_request.data_type == "interactions":
                return await self._export_interactions_data(export_request)
            elif export_request.data_type == "notifications":
                return await self._export_notifications_data(export_request)
            else:
                return []
                
        except Exception as e:
            logger.error("Error getting export data", 
                        data_type=export_request.data_type, 
                        error=str(e))
            return []
    
    async def _export_users_data(self, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Exportar datos de usuarios"""
        try:
            from app.models.user import User
            
            query = self.db.query(User).filter(User.site_id == export_request.site_id)
            
            # Aplicar filtros de fecha si existen
            if export_request.start_date:
                query = query.filter(User.fecha_registro >= export_request.start_date)
            if export_request.end_date:
                query = query.filter(User.fecha_registro <= export_request.end_date)
            
            users = query.all()
            
            export_data = []
            for user in users:
                user_data = {
                    "id": user.id,
                    "nombre": user.nombre,
                    "email": user.email,
                    "telefono": user.telefono,
                    "fecha_registro": user.fecha_registro.isoformat() if user.fecha_registro else None,
                    "puntos_acumulados": user.puntos_acumulados,
                    "instagram_seguido": user.instagram_seguido,
                    "reseñas_dejadas": user.reseñas_dejadas,
                    "videos_completados": user.videos_completados,
                    "total_descuento": user.total_descuento,
                    "activo": user.activo
                }
                
                # Filtrar campos específicos si se solicitan
                if export_request.fields:
                    user_data = {k: v for k, v in user_data.items() if k in export_request.fields}
                
                export_data.append(user_data)
            
            return export_data
            
        except Exception as e:
            logger.error("Error exporting users data", error=str(e))
            return []
    
    async def _export_stickers_data(self, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Exportar datos de stickers"""
        try:
            from app.models.sticker import Sticker
            
            query = self.db.query(Sticker).filter(Sticker.site_id == export_request.site_id)
            
            # Aplicar filtros de fecha si existen
            if export_request.start_date:
                query = query.filter(Sticker.fecha_generacion >= export_request.start_date)
            if export_request.end_date:
                query = query.filter(Sticker.fecha_generacion <= export_request.end_date)
            
            stickers = query.all()
            
            export_data = []
            for sticker in stickers:
                sticker_data = {
                    "id": sticker.id,
                    "usuario_id": sticker.usuario_id,
                    "tipo_sticker": sticker.tipo_sticker.value,
                    "codigo_descuento": sticker.codigo_descuento,
                    "porcentaje_descuento": sticker.porcentaje_descuento,
                    "fecha_generacion": sticker.fecha_generacion.isoformat() if sticker.fecha_generacion else None,
                    "fecha_expiracion": sticker.fecha_expiracion.isoformat() if sticker.fecha_expiracion else None,
                    "usado": sticker.usado,
                    "fecha_uso": sticker.fecha_uso.isoformat() if sticker.fecha_uso else None
                }
                
                # Filtrar campos específicos si se solicitan
                if export_request.fields:
                    sticker_data = {k: v for k, v in sticker_data.items() if k in export_request.fields}
                
                export_data.append(sticker_data)
            
            return export_data
            
        except Exception as e:
            logger.error("Error exporting stickers data", error=str(e))
            return []
    
    async def _export_interactions_data(self, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Exportar datos de interacciones"""
        try:
            from app.models.interaction import Interaction
            
            query = self.db.query(Interaction).filter(Interaction.site_id == export_request.site_id)
            
            # Aplicar filtros de fecha si existen
            if export_request.start_date:
                query = query.filter(Interaction.fecha_interaccion >= export_request.start_date)
            if export_request.end_date:
                query = query.filter(Interaction.fecha_interaccion <= export_request.end_date)
            
            interactions = query.all()
            
            export_data = []
            for interaction in interactions:
                interaction_data = {
                    "id": interaction.id,
                    "usuario_id": interaction.usuario_id,
                    "tipo_interaccion": interaction.tipo_interaccion.value,
                    "contenido_id": interaction.contenido_id,
                    "contenido_tipo": interaction.contenido_tipo,
                    "contenido": interaction.contenido,
                    "puntos_obtenidos": interaction.puntos_obtenidos,
                    "fecha_interaccion": interaction.fecha_interaccion.isoformat() if interaction.fecha_interaccion else None
                }
                
                # Filtrar campos específicos si se solicitan
                if export_request.fields:
                    interaction_data = {k: v for k, v in interaction_data.items() if k in export_request.fields}
                
                export_data.append(interaction_data)
            
            return export_data
            
        except Exception as e:
            logger.error("Error exporting interactions data", error=str(e))
            return []
    
    async def _export_notifications_data(self, export_request: ExportRequest) -> List[Dict[str, Any]]:
        """Exportar datos de notificaciones"""
        try:
            from app.models.notification import Notification
            
            query = self.db.query(Notification).filter(Notification.site_id == export_request.site_id)
            
            # Aplicar filtros de fecha si existen
            if export_request.start_date:
                query = query.filter(Notification.created_at >= export_request.start_date)
            if export_request.end_date:
                query = query.filter(Notification.created_at <= export_request.end_date)
            
            notifications = query.all()
            
            export_data = []
            for notification in notifications:
                notification_data = {
                    "id": notification.id,
                    "user_id": notification.user_id,
                    "type": notification.type.value,
                    "title": notification.title,
                    "message": notification.message,
                    "priority": notification.priority.value,
                    "status": notification.status.value,
                    "created_at": notification.created_at.isoformat() if notification.created_at else None,
                    "read_at": notification.read_at.isoformat() if notification.read_at else None,
                    "delivered_at": notification.delivered_at.isoformat() if notification.delivered_at else None
                }
                
                # Filtrar campos específicos si se solicitan
                if export_request.fields:
                    notification_data = {k: v for k, v in notification_data.items() if k in export_request.fields}
                
                export_data.append(notification_data)
            
            return export_data
            
        except Exception as e:
            logger.error("Error exporting notifications data", error=str(e))
            return []
    
    async def _generate_export_file(
        self, 
        export_id: str, 
        export_request: ExportRequest, 
        export_data: List[Dict[str, Any]]
    ) -> Path:
        """Generar archivo de exportación"""
        try:
            exports_dir = Path("exports")
            exports_dir.mkdir(exist_ok=True)
            
            file_path = exports_dir / f"{export_id}.{export_request.format.value}"
            
            if export_request.format == ReportFormat.CSV:
                await self._generate_csv_export(file_path, export_data, export_request.include_headers)
            elif export_request.format == ReportFormat.JSON:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            elif export_request.format == ReportFormat.XLSX:
                await self._generate_xlsx_export(file_path, export_data, export_request.include_headers)
            
            return file_path
            
        except Exception as e:
            logger.error("Error generating export file", 
                        export_id=export_id, 
                        format=export_request.format.value, 
                        error=str(e))
            raise
    
    async def _generate_csv_export(self, file_path: Path, export_data: List[Dict[str, Any]], include_headers: bool):
        """Generar exportación CSV"""
        try:
            if not export_data:
                return
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                fieldnames = export_data[0].keys()
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                
                if include_headers:
                    writer.writeheader()
                
                for row in export_data:
                    writer.writerow(row)
                    
        except Exception as e:
            logger.error("Error generating CSV export", error=str(e))
            raise
    
    async def _generate_xlsx_export(self, file_path: Path, export_data: List[Dict[str, Any]], include_headers: bool):
        """Generar exportación XLSX"""
        try:
            if not export_data:
                return
            
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            
            if include_headers and export_data:
                # Escribir encabezados
                headers = list(export_data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Escribir datos
                for row, data in enumerate(export_data, 2):
                    for col, value in enumerate(data.values(), 1):
                        ws.cell(row=row, column=col, value=value)
            else:
                # Escribir solo datos
                for row, data in enumerate(export_data, 1):
                    for col, value in enumerate(data.values(), 1):
                        ws.cell(row=row, column=col, value=value)
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error("Error generating XLSX export", error=str(e))
            raise
